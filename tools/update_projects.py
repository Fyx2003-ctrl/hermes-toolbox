#!/usr/bin/env python3
"""GitHub 开源项目定期检索更新脚本
每周运行: 按分类搜索热门项目, 合并进 github_projects.xlsx
- 已有项目保留状态列 (已复现/未复现/待验证)
- 新增项目自动追加为"待验证"
用法: python3 update_projects.py
"""
import json
import os
import sys
import time
import urllib.parse
import urllib.request

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(BASE_DIR, 'github_projects.xlsx')

QUERIES = [
    ('AI 助手框架', 'AI agent framework language:Python'),
    ('本地大模型', 'local LLM inference'),
    ('数据可视化', 'data visualization python'),
    ('粒子系统', 'particle system canvas'),
    ('文献计量', 'bibliometrics OR scientometrics'),
    ('知识库', 'personal knowledge base'),
]
PER_CAT = 8


def gh_search(query, per_page=PER_CAT):
    url = (f"https://api.github.com/search/repositories?q="
           f"{urllib.parse.quote(query)}&sort=stars&order=desc&per_page={per_page}")
    headers = {'User-Agent': 'hermes-agent', 'Accept': 'application/vnd.github+json'}
    token = os.environ.get('GH_TOKEN') or os.environ.get('GITHUB_TOKEN')
    if token:
        headers['Authorization'] = f'token {token}'
    req = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.load(r)


def fetch_projects():
    results = []
    for cat, q in QUERIES:
        try:
            d = gh_search(q)
            for item in d.get('items', [])[:PER_CAT]:
                results.append({
                    '分类': cat,
                    '项目': item['full_name'],
                    '星数': item['stargazers_count'],
                    '语言': item.get('language') or '-',
                    '简介': (item.get('description') or '')[:100],
                    'URL': item['html_url'],
                })
            print(f"  {cat}: {len(d.get('items', []))} 个")
        except Exception as e:
            print(f"  {cat}: 失败 {e}")
        time.sleep(1.2)
    return results


def style_sheet(ws):
    header_fill = PatternFill('solid', fgColor='2F5496')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    cat_fill = PatternFill('solid', fgColor='D6E4F0')
    done_fill = PatternFill('solid', fgColor='C6EFCE')
    todo_fill = PatternFill('solid', fgColor='FFEB9C')
    fail_fill = PatternFill('solid', fgColor='FFC7CE')
    thin = Side(style='thin', color='B0B0B0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['序号', '分类', '项目名', '⭐星数', '语言', '简介', '状态', '备注']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    row = 2
    current_cat = None
    idx = 0
    for p in ws._rows if False else []:
        pass
    return row


def rebuild_xlsx(projects, status_map):
    wb = Workbook()
    ws = wb.active
    ws.title = "开源项目汇总"

    header_fill = PatternFill('solid', fgColor='2F5496')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    cat_fill = PatternFill('solid', fgColor='D6E4F0')
    done_fill = PatternFill('solid', fgColor='C6EFCE')
    todo_fill = PatternFill('solid', fgColor='FFEB9C')
    fail_fill = PatternFill('solid', fgColor='FFC7CE')
    thin = Side(style='thin', color='B0B0B0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['序号', '分类', '项目名', '⭐星数', '语言', '简介', '状态', '备注']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    row = 2
    current_cat = None
    idx = 0
    for p in projects:
        if p['分类'] != current_cat:
            current_cat = p['分类']
            ws.cell(row, 1, f"【{current_cat}】").font = Font(bold=True)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            for c in range(1, 9):
                ws.cell(row, c).fill = cat_fill
                ws.cell(row, c).border = border
            row += 1
        idx += 1
        ws.cell(row, 1, idx)
        ws.cell(row, 2, p['分类'])
        ws.cell(row, 3, p['项目'])
        ws.cell(row, 4, p['星数'])
        ws.cell(row, 5, p['语言'])
        ws.cell(row, 6, p['简介'])
        status = status_map.get(p['项目'], '待验证')
        sc = ws.cell(row, 7, status)
        sc.fill = done_fill if status == '已复现' else (fail_fill if status == '未复现' else todo_fill)
        ws.cell(row, 8, '')
        for c in range(1, 9):
            ws.cell(row, c).border = border
            if c != 1:
                ws.cell(row, c).alignment = Alignment(vertical='center')
        row += 1

    widths = [6, 12, 30, 9, 10, 55, 10, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    ws2 = wb.create_sheet("说明")
    ws2['A1'] = '开源项目汇总表 - 使用说明'
    ws2['A1'].font = Font(bold=True, size=13)
    for r, t in enumerate([
        '1. 状态列含义:',
        '   已复现 = 已在本地运行/复现成功',
        '   未复现 = 尝试过但未成功',
        '   待验证 = 尚未在本地尝试',
        '',
        '2. 本表由定期脚本自动更新 (每周), 新增项目自动追加',
        f'3. 最近更新: {time.strftime("%Y-%m-%d %H:%M")}',
    ], 3):
        ws2.cell(r, 1, t)

    wb.save(XLSX_PATH)


def main():
    print(f"[1/2] 检索 GitHub 热门项目 ({time.strftime('%Y-%m-%d')})...")
    projects = fetch_projects()
    print(f"  共检索到 {len(projects)} 个项目")

    # 内置本地成果条目 (保证始终保留)
    LOCAL_ITEM = {
        '分类': '粒子系统',
        '项目': '★本地成果: 二次元粒子系统 (Hermes定制)',
        '星数': '-', '语言': 'Python/JS',
        '简介': '粒子汇聚成二次元角色(全彩/简笔画/线条模式, 8万粒子, 60fps)',
        'URL': '',
    }
    projects.append(LOCAL_ITEM)

    # 读取旧表保留状态
    status_map = {}
    if os.path.exists(XLSX_PATH):
        try:
            wb_old = load_workbook(XLSX_PATH)
            ws_old = wb_old['开源项目汇总']
            for row in ws_old.iter_rows(min_row=2, values_only=True):
                if row[2] and row[6] and row[6] in ('已复现', '未复现', '待验证'):
                    status_map[row[2]] = row[6]
        except Exception as e:
            print(f"  读取旧表失败: {e}")

    status_map[LOCAL_ITEM['项目']] = '已复现'
    new_cnt = sum(1 for p in projects if p['项目'] not in status_map)
    print(f"[2/2] 合并更新: 保留 {len(status_map)} 个历史项目, 新增 {new_cnt} 个")

    rebuild_xlsx(projects, status_map)
    print(f"✅ 已更新 {XLSX_PATH}")
    print(f"   汇总: {len(projects)} 个项目 | 新增 {new_cnt} 个 | 已复现 "
          f"{sum(1 for s in status_map.values() if s=='已复现')} 个")


if __name__ == '__main__':
    main()
